# web_ui.py
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    Response,
    send_file
)

from .security_utils import login_required, role_required
from .models import User, Attendance, Face, Notification, Complaint
from .database import db
from .security_utils import cipher
from datetime import datetime
from functools import wraps
from openpyxl import Workbook
import io
from sqlalchemy import extract, func
import calendar
web_ui = Blueprint("web_ui", __name__, template_folder="templates")


# DASHBOARD CHUNG
@web_ui.route("/dashboard")
@login_required
def dashboard():
    """Chuyển hướng dashboard theo role."""
    if session.get("role") == "admin":
        return redirect(url_for("web_ui.admin_dashboard"))
    else:
        return redirect(url_for("web_ui.employee_dashboard"))
@web_ui.route("/admin/dashboard")
@login_required
@role_required("admin")
def admin_dashboard():
    from datetime import datetime, time
    from sqlalchemy import extract, func
    import calendar

    now = datetime.now()
    today = now.date()
    month = now.month
    year = now.year

    total_users = User.query.count()
    total_logs = Attendance.query.count()
    recent = Attendance.query.order_by(Attendance.timestamp.desc()).limit(10).all()

    recent = Attendance.query.order_by(Attendance.timestamp.desc()).limit(10).all()

    for r in recent:
        r.work_hours = None 
        if r.action == "checkout":
            checkin = Attendance.query.filter(
                Attendance.user_id == r.user_id,
                Attendance.action == "checkin",
                Attendance.timestamp < r.timestamp
            ).order_by(Attendance.timestamp.desc()).first()

            if checkin:
                seconds = (r.timestamp - checkin.timestamp).total_seconds()
                r.work_hours = round(seconds / 3600, 2)

    total_seconds_today = 0
    working_users = set()

    users = User.query.filter_by(role="employee").all()

    for u in users:
        logs = Attendance.query.filter(
            Attendance.user_id == u.id,
            func.date(Attendance.timestamp) == today
        ).order_by(Attendance.timestamp.asc()).all()

        checkin_time = None

        for log in logs:
            if log.action == "checkin":
                checkin_time = log.timestamp
            elif log.action == "checkout" and checkin_time:
                duration = (log.timestamp - checkin_time).total_seconds()
                total_seconds_today += duration
                working_users.add(u.id)
                checkin_time = None 

    total_hours_today = round(total_seconds_today / 3600, 2)
    avg_hours_per_employee = (
        round(total_hours_today / len(working_users), 2)
        if working_users else 0
    )
    late_today = 0
    late_limit = time(8, 0)

    for u in users:
        first_checkin = Attendance.query.filter(
            Attendance.user_id == u.id,
            Attendance.action == "checkin",
            func.date(Attendance.timestamp) == today
        ).order_by(Attendance.timestamp.asc()).first()

        if first_checkin and first_checkin.timestamp.time() > late_limit:
            late_today += 1
    days_in_month = calendar.monthrange(year, month)[1]
    chart_days = list(range(1, days_in_month + 1))
    chart_values = [0] * days_in_month

    stats = (
        db.session.query(
            extract("day", Attendance.timestamp).label("day"),
            func.count(Attendance.id)
        )
        .filter(
            Attendance.action == "checkin",
            extract("month", Attendance.timestamp) == month,
            extract("year", Attendance.timestamp) == year
        )
        .group_by("day")
        .all()
    )

    for day, count in stats:
        chart_values[int(day) - 1] = count
    users_for_select = sorted(
        User.query.all(),
        key=lambda u: (u.full_name or "").lower()
    )

    months = [{"value": i, "label": f"Tháng {i}"} for i in range(1, 13)]

    return render_template(
        "admin_dashboard.html",
        total_users=total_users,
        total_logs=total_logs,
        total_hours_today=total_hours_today,
        avg_hours_per_employee=avg_hours_per_employee,
        late_today=late_today,
        recent=recent,
        users_for_select=users_for_select,
        months=months,
        current_month_num=month,
        chart_days=chart_days,
        chart_values=chart_values
    )



# EMPLOYEE DASHBOARD
@web_ui.route("/employee/dashboard")
@role_required("employee")
def employee_dashboard():
    user_id = session.get("user_id")
    user = User.query.get(user_id)

    recent = Attendance.query.filter_by(user_id=user_id)\
        .order_by(Attendance.timestamp.desc())\
        .limit(20).all()

    return render_template(
        "employee_dashboard.html",
        user=user,
        recent=recent
    )


# ADMIN: QUẢN LÝ USER
@web_ui.route("/admin/users")
@role_required("admin")
def admin_users():
    users = User.query.order_by(User.id.asc()).all()
    return render_template("admin_users.html", users=users)


@web_ui.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@role_required("admin")
def admin_users_delete(user_id):
    if user_id == session.get("user_id"):
        flash("Không thể tự xóa chính mình!", "warning")
        return redirect(url_for("web_ui.admin_users"))

    User.query.filter_by(id=user_id).delete()
    db.session.commit()

    flash("Đã xóa người dùng!", "success")
    return redirect(url_for("web_ui.admin_users"))


# XEM LỊCH SỬ CHẤM CÔNG
@web_ui.route("/attendance")
@login_required
def attendance_page():
    role = session.get("role")

    if role == "admin":
        logs = Attendance.query.order_by(Attendance.timestamp.desc()).limit(200).all()
    else:
        logs = Attendance.query.filter_by(user_id=session["user_id"])\
            .order_by(Attendance.timestamp.desc())\
            .limit(200).all()

    return render_template("attendance.html", logs=logs)


# ĐĂNG KÝ KHUÔN MẶT
@web_ui.route("/enroll_face", methods=["GET", "POST"])
@login_required
def enroll_face_page():
    if request.method == "POST":
        file = request.files.get("face_image")

        if not file:
            flash("Vui lòng chọn file!", "danger")
            return redirect(url_for("web_ui.enroll_face_page"))

        user_id = session.get("user_id")

        # Kiểm tra FK trước khi lưu
        user = User.query.get(user_id)
        if not user:
            flash("User không tồn tại trong hệ thống!", "danger")
            return redirect(url_for("web_ui.enroll_face_page"))

        img_bytes = file.read()
        encrypted = cipher.encrypt(img_bytes)

        face = Face(
            user_id=user_id,
            image_encrypted=encrypted
        )

        db.session.add(face)
        db.session.commit()

        flash("Đăng ký khuôn mặt thành công!", "success")
        return redirect(url_for("web_ui.employee_dashboard"))

    return render_template("enroll_face.html")


# ADMIN: TẠO NHÂN VIÊN MỚI
@web_ui.route("/admin/create_employee", methods=["GET", "POST"])
@role_required("admin")
def create_employee():
    if request.method == "POST":

        # Lấy dữ liệu form
        username = request.form.get("username")
        password = request.form.get("password")
        full_name = request.form.get("full_name")
        age = request.form.get("age")
        dob = request.form.get("dob")
        phone = request.form.get("phone")
        address = request.form.get("address")
        email = request.form.get("email")
        position = request.form.get("position")
        department = request.form.get("department")

        # Kiểm tra trùng username
        if User.query.filter_by(username=username).first():
            flash("Tên đăng nhập đã tồn tại!", "danger")
            return redirect(url_for("web_ui.create_employee"))

        new_id = generate_employee_id()

        user = User(
            id=new_id,               
            username=username,
            dob=dob,
            phone=phone,
            address=address,
            email=email,
            position=position,
            department=department,
            role="employee"
        )


        user.set_password(password)

        # Full name & Age cần dùng setter (do mã hóa Fernet)
        if full_name:
            user.full_name = full_name

        if age:
            user.age = int(age)

        db.session.add(user)
        db.session.commit()

        flash("Tạo nhân viên mới thành công!", "success")
        return redirect(url_for("web_ui.admin_dashboard"))

    return render_template("create_employee.html")


# EXPORT DỮ LIỆU CHẤM CÔNG
@web_ui.route("/export")
@role_required("admin")
def export_data():
    logs = Attendance.query.order_by(Attendance.timestamp.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header
    ws.append([
        "ID",
        "Mã nhân viên",
        "Tên nhân viên",
        "Hành động",
        "Thời gian"
    ])

    # Data
    for r in logs:
        ws.append([
            r.id,
            r.user_id,
            r.user.full_name or r.user.username,
            "Check-in" if r.action == "checkin" else "Check-out",
            r.timestamp.strftime("%d/%m/%Y %H:%M:%S")
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="bang_cham_cong.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@web_ui.route("/export/users")
@role_required("admin")
def export_users():
    from openpyxl import Workbook
    import io
    from flask import send_file

    users = User.query.order_by(User.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    ws.append([
        "ID",
        "Username",
        "Họ tên",
        "Email",
        "Phòng ban",
        "Chức vụ",
        "Role"
    ])

    for u in users:
        ws.append([
            u.id,
            u.username,
            u.full_name,
            u.email,
            u.department,
            u.position,
            u.role
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="danh_sach_nhan_vien.xlsx"
    )
@web_ui.route("/export/employee/<int:user_id>")
@role_required("admin")
def export_employee_attendance(user_id):
    from openpyxl import Workbook
    import io
    from flask import send_file

    user = User.query.get_or_404(user_id)
    logs = Attendance.query.filter_by(user_id=user_id)\
                           .order_by(Attendance.timestamp.asc())\
                           .all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append([
        "ID",
        "Hành động",
        "Thời gian"
    ])

    for r in logs:
        ws.append([
            r.id,
            "Check-in" if r.action == "checkin" else "Check-out",
            r.timestamp.strftime("%d/%m/%Y %H:%M:%S")
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"cham_cong_{user.username}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )

# ADMIN GỬI THÔNG BÁO
@web_ui.route("/send_notification", methods=["POST"])
@role_required("admin")
def send_notification():
    msg = request.form.get("message")
    if msg:
        notif = Notification(
            message=msg,
            sender_id=session["user_id"],
            receiver_id=None  
        )

        db.session.add(notif)
        db.session.commit()

        flash("Đã gửi thông báo!", "success")

    return redirect(url_for("web_ui.admin_dashboard"))


# EMPLOYEE GỬI KHIẾU NẠI
@web_ui.route("/complaint/<int:attendance_id>", methods=["POST"])
@role_required("employee")
def submit_complaint(attendance_id):
    reason = request.form.get("reason")

    if reason:
        c = Complaint(
            user_id=session["user_id"],
            attendance_id=attendance_id,
            reason=reason
        )
        db.session.add(c)
        
        #CREATE NOTIFICATION FOR ADMIN
        admins = User.query.filter_by(role="admin").all()

        for admin in admins:
            db.session.add(Notification(
                message=f"Khiếu nại mới từ {session.get('username')}",
                sender_id=session["user_id"],
                receiver_id=admin.id,
                is_read=False
            ))
        db.session.commit()
        flash("Đã gửi khiếu nại!", "success")

    return redirect(url_for("web_ui.employee_dashboard"))



# LOAD THÔNG BÁO CHO NHÂN VIÊN (GLOBAL)
@web_ui.app_context_processor
def inject_notifications():
    user_id = session.get("user_id")
    if not user_id:
        return dict(notifications=[])

    notifs = Notification.query.filter(
        (Notification.receiver_id == user_id) |
        (Notification.receiver_id == None)
    ).order_by(Notification.created_at.desc()).limit(5).all()

    return dict(notifications=notifs)

@web_ui.route("/admin/user_month_view")
@role_required("admin")
def user_month_view():
    from datetime import datetime
    import calendar
    from sqlalchemy import extract

    user_id = request.args.get("user_id", type=int)
    month = request.args.get("month", type=int)
    year = datetime.now().year

    if not user_id or not month:
        flash("Vui lòng chọn nhân viên và tháng!", "warning")
        return redirect(url_for("web_ui.admin_dashboard"))

    days_in_month = calendar.monthrange(year, month)[1]

    # 0 = chưa chấm, 1 = đã check-in, 2 = đủ check-in + check-out
    daily_status = {day: 0 for day in range(1, days_in_month + 1)}

    logs = Attendance.query.filter(
        Attendance.user_id == user_id,
        extract("month", Attendance.timestamp) == month,
        extract("year", Attendance.timestamp) == year
    ).all()

    for log in logs:
        day = log.timestamp.day

        if log.action == "checkin":
            daily_status[day] = max(daily_status[day], 1)
        elif log.action == "checkout":
            daily_status[day] = 2   

    user = User.query.get_or_404(user_id)

    return render_template(
        "admin_month_grid.html",
        user=user,
        month=month,
        daily_status=daily_status
    )

@web_ui.route("/checkout", methods=["POST"])
@login_required
def checkout():
    user_id = session.get("user_id")
    now = datetime.now()
    last_checkin = Attendance.query.filter_by(
        user_id=user_id,
        action="checkin"
    ).order_by(Attendance.timestamp.desc()).first()

    if not last_checkin:
        flash("Bạn chưa check-in!", "warning")
        return redirect(url_for("web_ui.employee_dashboard"))

    duration_hours = (now - last_checkin.timestamp).total_seconds() / 3600

    checkout_log = Attendance(
        user_id=user_id,
        action="checkout",
        timestamp=now,
        duration=round(duration_hours, 2)
    )

    db.session.add(checkout_log)
    db.session.commit()

    flash(f"Check-out thành công ({round(duration_hours,2)}h)", "success")
    return redirect(url_for("web_ui.employee_dashboard"))

def generate_employee_id():
    year = datetime.now().year
    prefix = str(year)

    # Lấy user id lớn nhất trong năm đó
    last = User.query.filter(User.id.like(f"{prefix}%")) \
                     .order_by(User.id.desc()) \
                     .first()
    if not last:
        return int(prefix + "0001")
    # tăng số cuối
    last_number = int(str(last.id)[4:])
    new_number = last_number + 1
    return int(f"{prefix}{new_number:04d}")
@web_ui.route("/admin/complaints")
@role_required("admin")
def admin_complaints():
    complaints = Complaint.query.order_by(Complaint.created_at.desc()).all()
    return render_template("admin_complaints.html", complaints=complaints)
@web_ui.route("/admin/complaints/<int:id>/update", methods=["POST"])
@role_required("admin")
def update_complaint_status(id):
    status = request.form.get("status")
    complaint = Complaint.query.get_or_404(id)

    if status not in ["approved", "rejected"]:
        flash("Trạng thái không hợp lệ", "danger")
        return redirect(url_for("web_ui.admin_complaints"))

    complaint.status = status

    # Gửi thông báo cho employee
    db.session.add(Notification(
        message=f"Khiếu nại của bạn đã được { 'duyệt' if status=='approved' else 'từ chối' }",
        sender_id=session["user_id"],
        receiver_id=complaint.user_id
    ))

    db.session.commit()
    flash("Đã cập nhật khiếu nại", "success")
    return redirect(url_for("web_ui.admin_complaints"))
